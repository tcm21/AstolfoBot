import openpyxl
from io import BytesIO
import urllib.request
import pandas as pd


def load_workbook_from_url(url):
    file = urllib.request.urlopen(url).read()
    return openpyxl.load_workbook(filename=BytesIO(file), data_only=True)


def get_drop_rates(item_name: str, region: str = "JP"):
    if region == "JP":
        url = "https://docs.google.com/spreadsheets/d/1_SlTjrVRTgHgfS7sRqx4CeJMqlz687HdSlYqiW-JvQA/edit#gid=843570146"
    else:
        url = "https://docs.google.com/spreadsheets/d/1_SlTjrVRTgHgfS7sRqx4CeJMqlz687HdSlYqiW-JvQA/edit#gid=1676231111"

    url = url.replace('/edit#gid=', '/export?format=xlsx&gid=')
    wb = load_workbook_from_url(url)
    ws = wb.active
    merged_cells = ws.merged_cells.ranges
    found_cells = find_in_cells(item_name, merged_cells)
    if len(found_cells) != 2: return None
    data_rows = []
    for row in ws.iter_rows(min_row=found_cells[0].row, max_row=found_cells[0].row + 4, min_col=found_cells[0].column + 1, max_col=found_cells[1].column - 1):
        data_row = []
        hyperlink_target = ""
        for cell in row:
            data_row.append(cell.value)
            if cell.hyperlink: hyperlink_target = cell.hyperlink.target
        data_row.append(hyperlink_target)
        data_rows.append(data_row)
        # data_rows.append([cell.value for cell in row])
    df = pd.DataFrame(data_rows)
    df = df.rename(columns={
        0: 'No.',
        1: 'Code',
        2: 'Area',
        3: 'Quest',
        4: 'AP',
        5: 'BP/AP',
        6: 'AP/Drop',
        7: 'AP Suffix',
        8: 'Drop chance',
        9: 'Drop chance Suffix',
        10: 'Runs',
        11: 'Hyperlink',
    })

    return df[df["Code"].notnull()]


def find_in_cells(item_name: str, merged_cells):
    found_cells = []
    for range in merged_cells:
        for cells in range.ws[range.coord]:
            for cell in cells:
                if cell.value and item_name.upper() in str(cell.value).upper():
                    found_cells.append(cell)
    
    return sorted(found_cells, key=lambda cell: cell.column)
